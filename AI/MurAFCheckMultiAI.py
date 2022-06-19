import numpy as np
import os
import pandas as pd
import librosa
import librosa.display as display
import glob 
import matplotlib.pyplot as plt
import tensorflow
from tensorflow.keras.models import load_model
import openpyxl as px
from openpyxl.styles import PatternFill


def extract_data(file_name):
    # function to load files and extract features
    try:
        # here kaiser_fast is a technique used for faster extraction
        X, sample_rate = librosa.load(file_name, res_type='kaiser_fast')
        # we extract mfcc feature from data
        global mfccs
        mfccs = np.mean(librosa.feature.mfcc(
            y=X, sr=sample_rate, n_mfcc=40).T, axis=0)
    except Exception as e:
        #Error Message
        print("Error encountered while parsing file: ")

    #Reshape data to be in the right matrix
    feature = np.array(mfccs).reshape([-1, 1])
    print(mfccs)
    return feature

af = load_model('AI/Models/AF.h5')#AF AI Path
murmur = load_model('AI/Models/murmur.h5') #murmur AI Path

#variables
mfccs = 0.0
data1 = []
data2 = []
data3 = []
data4 = []
data5 = []
data6 = []
AFList = []
MurList = []
returnvalue= []
ex = '25'
Sheet = 'Version2'
file_name1 = 'C:/Users/mattc/OneDrive/Documents/Project/CSM/Testing/AITest/A22/Test1.wav' #WAV File Path
file_name2 = 'C:/Users/mattc/OneDrive/Documents/Project/CSM/HeartDiseaseSounds/ReducNoise_Testing/TestFile1.wav'
file_name3 = 'C:/Users/mattc/OneDrive/Documents/Project/CSM/HeartDiseaseSounds/ReducNoise_Testing/TestFile2.wav'
file_name4 = 'C:/Users/mattc/OneDrive/Documents/Project/CSM/HeartDiseaseSounds/ReducNoise_Testing/TestFile3.wav'
file_name5 = 'C:/Users/mattc/OneDrive/Documents/Project/CSM/HeartDiseaseSounds/ReducNoise_Testing/TestFile4.wav'
file_name6 = 'C:/Users/mattc/OneDrive/Documents/Project/CSM/HeartDiseaseSounds/ReducNoise_Testing/TestFile5.wav'

ExcelFiles = 'C:/Users/mattc/OneDrive/Documents/Project/CSM/Testing/AITest/ExprimentResult.xlsx'#Excel File
#Extract and read the WAV file for the AI
data1.append(extract_data(file_name1))
data2.append(extract_data(file_name2))
data3.append(extract_data(file_name3))
data4.append(extract_data(file_name4))
data5.append(extract_data(file_name5))
data6.append(extract_data(file_name6))

#columns values
col1 = 'B'
col2 = 'C'
col3 = 'D'
col4 = 'E'
col5 = 'F'
col6 = 'G'
col7 = 'H'
col8 = 'I'
col9 = 'J'
col10 = 'K'
col11 = 'L'
col12 = 'M'

#Analize Predict and Save the data
af_result = af.predict(np.array(data1))
murmur_result = murmur.predict(np.array(data1))
y1 = af_result[0]
b1 = murmur_result[0]

af_result = af.predict(np.array(data2))
murmur_result = murmur.predict(np.array(data2))
y2 = af_result[0]
b2 = murmur_result[0]

af_result = af.predict(np.array(data3))
murmur_result = murmur.predict(np.array(data3))
y3 = af_result[0]
b3 = murmur_result[0]

af_result = af.predict(np.array(data4))
murmur_result = murmur.predict(np.array(data4))
y4 = af_result[0]
b4 = murmur_result[0]

af_result = af.predict(np.array(data5))
murmur_result = murmur.predict(np.array(data5))
y5 = af_result[0]
b5 = murmur_result[0]

af_result = af.predict(np.array(data6))
murmur_result = murmur.predict(np.array(data6))
y6 = af_result[0]
b6 = murmur_result[0]

#Make it in % instead of 0.
af_return = y1[0]*100
murmur_return = b1[0]*100

#Show the ai confidence[AF,murmur]
returnvalue.append([af_return, murmur_return])
print('Normal: ',returnvalue[0])

af_return = y2[0]*100
murmur_return = b2[0]*100
returnvalue.append([af_return, murmur_return])
print('50hz.: ',returnvalue[1])

af_return = y3[0]*100
murmur_return = b3[0]*100
returnvalue.append([af_return, murmur_return])
print('100hz.: ',returnvalue[2])


af_return = y4[0]*100
murmur_return = b4[0]*100
returnvalue.append([af_return, murmur_return])
print('200hz.: ',returnvalue[3])

af_return = y5[0]*100
murmur_return = b5[0]*100
returnvalue.append([af_return, murmur_return])
print('300hz.: ',returnvalue[4])

af_return = y6[0]*100
murmur_return = b6[0]*100
returnvalue.append([af_return, murmur_return])
print('400hz.: ',returnvalue[5])

#Upload to Excel
wb = px.load_workbook(ExcelFiles)
wsbook = wb.active
ws = wb[Sheet]

#Input the value to Excel
print('Changing Value of Excel Cell')
ws[col1+ex].value = round(returnvalue[0][0],3)
ws[col2+ex].value = round(returnvalue[0][1],3)
ws[col3+ex].value = round(returnvalue[1][0],3)
ws[col4+ex].value = round(returnvalue[1][1],3)
ws[col5+ex].value = round(returnvalue[2][0],3)
ws[col6+ex].value = round(returnvalue[2][1],3) 
ws[col7+ex].value = round(returnvalue[3][0],3)
ws[col8+ex].value = round(returnvalue[3][1],3)
ws[col9+ex].value = round(returnvalue[4][0],3)
ws[col10+ex].value = round(returnvalue[4][1],3)
ws[col11+ex].value = round(returnvalue[5][0],3)
ws[col12+ex].value = round(returnvalue[5][1],3)
print('Value of Excel Cell changed')

CopiedValue = returnvalue.copy()
#Find min value in AF and Murmur
CopiedValue.sort()
AFMin = CopiedValue[0][0]
print(AFMin)
AFMax = CopiedValue[len(CopiedValue)-1][0]
print(AFMax)
for i in range (len(CopiedValue)):
    CopiedValue[i].reverse()
CopiedValue.sort()
MurMin = CopiedValue[0][0]
MurMax = CopiedValue[len(CopiedValue)-1][0]

#Colour the cell in excel
print('Changing Colour of Excel Cell')
Red = PatternFill(patternType='solid', fgColor='35FC03')
Green = PatternFill(patternType='solid', fgColor='FC2C03')

if returnvalue[0][1] == AFMax:
    ws[col1+ex].fill = Green
if returnvalue[1][1] == AFMax:
    ws[col3+ex].fill = Green
if returnvalue[2][1] == AFMax:
    ws[col5+ex].fill = Green
if returnvalue[3][1] == AFMax:
    ws[col7+ex].fill = Green
if returnvalue[4][1] == AFMax:
    ws[col9+ex].fill = Green
if returnvalue[5][1] == AFMax:
    ws[col11+ex].fill = Green

if returnvalue[0][0] == MurMax:
    ws[col2+ex].fill = Green
if returnvalue[1][0] == MurMax:
    ws[col4+ex].fill = Green
if returnvalue[2][0] == MurMax:
    ws[col6+ex].fill = Green
if returnvalue[3][0] == MurMax:
    ws[col8+ex].fill = Green
if returnvalue[4][0] == MurMax:
    ws[col10+ex].fill= Green
if returnvalue[5][0] == MurMax:
    ws[col12+ex].fill = Green



if returnvalue[0][1] == AFMin:
    ws[col1+ex].fill = Red
if returnvalue[1][1] == AFMin:
    ws[col3+ex].fill = Red
if returnvalue[2][1] == AFMin:
    ws[col5+ex].fill = Red
if returnvalue[3][1] == AFMin:
    ws[col7+ex].fill = Red
if returnvalue[4][1] == AFMin:
    ws[col9+ex].fill = Red
if returnvalue[5][1] == AFMin:
    ws[col11+ex].fill = Red

if returnvalue[0][0] == MurMin:
    ws[col2+ex].fill = Red
if returnvalue[1][0] == MurMin:
    ws[col4+ex].fill = Red
if returnvalue[2][0] == MurMin:
    ws[col6+ex].fill = Red
if returnvalue[3][0] == MurMin:
    ws[col8+ex].fill = Red
if returnvalue[4][0] == MurMin:
    ws[col10+ex].fill= Red
if returnvalue[5][0] == MurMin:
    ws[col12+ex].fill = Red

print('Excel colour changed')

#Save To excel
wb.save(ExcelFiles)
print('Imported to Excel')

