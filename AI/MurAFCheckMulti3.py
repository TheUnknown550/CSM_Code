import numpy as np
import os
import pandas as pd
import librosa
import librosa.display as display
import glob 
import matplotlib.pyplot as plt
import tensorflow
from tensorflow.keras.models import load_model
import openpyxl as px

#***variables***
mfccs = 0.0
result = []
col = 6
row1 = 'D'
row2 = 'E'
file = 'C:/Users/mattc/OneDrive/Documents/Project/CSM/Testing/TubeTest/Moto/7inch'
#file = 'C:/Users/mattc/OneDrive/Documents/Project/CSM/Testing/DACTest/A22/EZRA/Ezra'#WAV File Path

#Make list for all files
file_name1 = file+'1.wav'
file_name2 = file+'2.wav'
file_name3 = file+'3.wav'

ExcelFile = 'C:/Users/mattc/OneDrive/Documents/Project/CSM/Testing/TubeTest/TubeResult.xlsx' #Excel File Path  .xlsx
Sheet = 'Moto' #Excel Sheet that is being edited

data = []
data1 = []
data2 = []
data3 = []
returnvalue = []

#Upload to Excel
wb = px.load_workbook(ExcelFile)
wsbook = wb.active
ws = wb[Sheet]

#Function to extract file
def extract_data(file_name):
    # function to load files and extract features
    try:
        # here kaiser_fast is a technique used for faster extraction
        X, sample_rate = librosa.load(file_name, res_type='kaiser_fast')
        # we extract mfcc feature from data
        global mfccs
        mfccs = np.mean(librosa.feature.mfcc(
            y=X, sr=sample_rate, n_mfcc=40).T, axis=0)
    except Exception as e:
        #Error Message
        print("Error encountered while parsing file: ")
    #Reshape data to be in the right matrix
    feature = np.array(mfccs).reshape([-1, 1])
    print(mfccs)
    return feature


af = load_model('AI/Models/AF.h5')#AF AI Path
murmur = load_model('AI/Models/murmur.h5') #murmur AI Path


#Extract and read the WAV file for the AI
a = extract_data(file_name1)
data.append(a)

#Analize Predict and Save the data
af_result = af.predict(np.array(data))
murmur_result = murmur.predict(np.array(data))
y = af_result[0]
b = murmur_result[0]

#Make it in % instead of 0.
af_return = y[0]*100
murmur_return = b[0]*100

#Show the ai confidence[AF,murmur]
returnvalue.append([af_return, murmur_return])

#Do the same for all files
a = extract_data(file_name2)
data1.append(a)
af_result = af.predict(np.array(data1))
murmur_result = murmur.predict(np.array(data1))
y = af_result[0]
b = murmur_result[0]
af_return = y[0]*100
murmur_return = b[0]*100
returnvalue.append([af_return, murmur_return])


a = extract_data(file_name3)
data2.append(a)
af_result = af.predict(np.array(data2))
murmur_result = murmur.predict(np.array(data2))
y = af_result[0]
b = murmur_result[0]
af_return = y[0]*100
murmur_return = b[0]*100
returnvalue.append([af_return, murmur_return])


#Show AI output in terminal
print('File1: ',returnvalue[0])
print('File2: ',returnvalue[1])
print('File3: ',returnvalue[2])



#Input the value to Excel
ws[row1+str(col)].value = round(returnvalue[0][0],3)
ws[row2+str(col)].value = round(returnvalue[0][1],3)
ws[row1+str(col+1)].value = round(returnvalue[1][0],3)
ws[row2+str(col+1)].value = round(returnvalue[1][1],3)
ws[row1+str(col+2)].value = round(returnvalue[2][0],3)
ws[row2+str(col+2)].value = round(returnvalue[2][1],3)

#Save To excel
wb.save(ExcelFile)
print('\nImported to Excel')