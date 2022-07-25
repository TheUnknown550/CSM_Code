import numpy as np
import os
import pandas as pd
import librosa
import librosa.display as display
import glob 
import matplotlib.pyplot as plt
import tensorflow
from tensorflow.keras.models import load_model
import openpyxl as px
from openpyxl.styles import PatternFill

#***variables***
mfccs = 0.0
result = []
col = 6
row1 = 'F'
row2 = 'G'
file = 'C:/Users/mattc/OneDrive/Documents/Project/CSM/Testing/DACTest/A22/HIFI/Hifi'#WAV File Path

#Make list for all files
file_name = [file+'Test1.wav', file+'Test2.wav', file+'Test3.wav',
file+'Test4.wav', file+'Test5.wav', file+'Test6.wav', file+'Test7.wav',
file+'Test8.wav', file+'Test9.wav', file+'Test10.wav', file+'Test11.wav']

ExcelFile = 'C:/Users/mattc/OneDrive/Documents/Project/CSM/Testing/DACTest/ExprimentResultDAC.xlsx' #Excel File Path  .xlsx
Sheet = 'A22' #Excel Sheet that is being edited

data = []
data1 = []
data2 = []
data3 = []
data4 = []
data5 = []
data6 = []
data7 = []
data8 = []
data9 = []
data10 = []
data11 = []
returnvalue = []

#Upload to Excel
wb = px.load_workbook(ExcelFile)
wsbook = wb.active
ws = wb[Sheet]

#Function to extract file
def extract_data(file_name):
    # function to load files and extract features
    try:
        # here kaiser_fast is a technique used for faster extraction
        X, sample_rate = librosa.load(file_name, res_type='kaiser_fast')
        # we extract mfcc feature from data
        global mfccs
        mfccs = np.mean(librosa.feature.mfcc(
            y=X, sr=sample_rate, n_mfcc=40).T, axis=0)
    except Exception as e:
        #Error Message
        print("Error encountered while parsing file: ")
    #Reshape data to be in the right matrix
    feature = np.array(mfccs).reshape([-1, 1])
    print(mfccs)
    return feature

af = load_model('AI/Models/AF.h5')#AF AI Path
murmur = load_model('AI/Models/murmur.h5') #murmur AI Path


#Extract and read the WAV file for the AI
a = extract_data(file_name[0])
data.append(a)

#Analize Predict and Save the data
af_result = af.predict(np.array(data))
murmur_result = murmur.predict(np.array(data))
y = af_result[0]
b = murmur_result[0]

#Make it in % instead of 0.
af_return = y[0]*100
murmur_return = b[0]*100

#Show the ai confidence[AF,murmur]
returnvalue.append([af_return, murmur_return])

#Do the same for all files
a = extract_data(file_name[1])
data1.append(a)
af_result = af.predict(np.array(data1))
murmur_result = murmur.predict(np.array(data1))
y = af_result[0]
b = murmur_result[0]
af_return = y[0]*100
murmur_return = b[0]*100
returnvalue.append([af_return, murmur_return])


a = extract_data(file_name[2])
data2.append(a)
af_result = af.predict(np.array(data2))
murmur_result = murmur.predict(np.array(data2))
y = af_result[0]
b = murmur_result[0]
af_return = y[0]*100
murmur_return = b[0]*100
returnvalue.append([af_return, murmur_return])


a = extract_data(file_name[3])
data3.append(a)
af_result = af.predict(np.array(data3))
murmur_result = murmur.predict(np.array(data3))
y = af_result[0]
b = murmur_result[0]
af_return = y[0]*100
murmur_return = b[0]*100
returnvalue.append([af_return, murmur_return])


a = extract_data(file_name[4])
data4.append(a)
af_result = af.predict(np.array(data4))
murmur_result = murmur.predict(np.array(data4))
y = af_result[0]
b = murmur_result[0]
af_return = y[0]*100
murmur_return = b[0]*100
returnvalue.append([af_return, murmur_return])


a = extract_data(file_name[5])
data5.append(a)
af_result = af.predict(np.array(data5))
murmur_result = murmur.predict(np.array(data5))
y = af_result[0]
b = murmur_result[0]
af_return = y[0]*100
murmur_return = b[0]*100
returnvalue.append([af_return, murmur_return])


a = extract_data(file_name[6])
data6.append(a)
af_result = af.predict(np.array(data6))
murmur_result = murmur.predict(np.array(data6))
y = af_result[0]
b = murmur_result[0]
af_return = y[0]*100
murmur_return = b[0]*100
returnvalue.append([af_return, murmur_return])


a = extract_data(file_name[7])
data7.append(a)
af_result = af.predict(np.array(data7))
murmur_result = murmur.predict(np.array(data7))
y = af_result[0]
b = murmur_result[0]
af_return = y[0]*100
murmur_return = b[0]*100
returnvalue.append([af_return, murmur_return])


a = extract_data(file_name[8])
data8.append(a)
af_result = af.predict(np.array(data8))
murmur_result = murmur.predict(np.array(data8))
y = af_result[0]
b = murmur_result[0]
af_return = y[0]*100
murmur_return = b[0]*100
returnvalue.append([af_return, murmur_return])


a = extract_data(file_name[9])
data9.append(a)
af_result = af.predict(np.array(data9))
murmur_result = murmur.predict(np.array(data9))
y = af_result[0]
b = murmur_result[0]
af_return = y[0]*100
murmur_return = b[0]*100
returnvalue.append([af_return, murmur_return])


a = extract_data(file_name[10])
data10.append(a)
af_result = af.predict(np.array(data10))
murmur_result = murmur.predict(np.array(data10))
y = af_result[0]
b = murmur_result[0]
af_return = y[0]*100
murmur_return = b[0]*100
returnvalue.append([af_return, murmur_return])
print(returnvalue[0][0])

#Show AI output in terminal
print('File1: ',returnvalue[0])
print('File2: ',returnvalue[1])
print('File3: ',returnvalue[2])
print('File4: ',returnvalue[3])
print('File5: ',returnvalue[4])
print('File6: ',returnvalue[5])
print('File7: ',returnvalue[6])
print('File8: ',returnvalue[7])
print('File9: ',returnvalue[8])
print('File10: ',returnvalue[9])
print('File11: ',returnvalue[10])

'''
#Input the value to Excel
ws[row1+str(col)].value = round(returnvalue[0][0],3)
ws[row2+str(col)].value = round(returnvalue[0][1],3)
ws[row1+str(col+1)].value = round(returnvalue[1][0],3)
ws[row2+str(col+1)].value = round(returnvalue[1][1],3)
ws[row1+str(col+2)].value = round(returnvalue[2][0],3)
ws[row2+str(col+2)].value = round(returnvalue[2][1],3)
ws[row1+str(col+3)].value = round(returnvalue[3][0],3)
ws[row2+str(col+3)].value = round(returnvalue[3][1],3) 
ws[row1+str(col+4)].value = round(returnvalue[4][0],3)
ws[row2+str(col+4)].value = round(returnvalue[4][1],3)
ws[row1+str(col+5)].value = round(returnvalue[5][0],3)
ws[row2+str(col+5)].value = round(returnvalue[5][1],3)
ws[row1+str(col+6)].value = round(returnvalue[6][0],3)
ws[row2+str(col+6)].value = round(returnvalue[6][1],3)
ws[row1+str(col+7)].value = round(returnvalue[7][0],3)
ws[row2+str(col+7)].value = round(returnvalue[7][1],3)
ws[row1+str(col+8)].value = round(returnvalue[8][0],3)
ws[row2+str(col+8)].value = round(returnvalue[8][1],3)
ws[row1+str(col+9)].value = round(returnvalue[9][0],3)
ws[row2+str(col+9)].value = round(returnvalue[9][1],3)
ws[row1+str(col+10)].value = round(returnvalue[10][0],3)
ws[row2+str(col+10)].value = round(returnvalue[10][1],3)


CopiedValue = returnvalue.copy()
#Find min value in AF and Murmur
CopiedValue.sort()
AFMin = CopiedValue[0][0]
AFMax = CopiedValue[len(CopiedValue)-1][0]
for i in range (len(CopiedValue)):
    CopiedValue[i].reverse()
CopiedValue.sort()
MurMin = CopiedValue[0][0]
MurMax = CopiedValue[len(CopiedValue)-1][0]

#Colour the cell in excel
print('Changing Colour of Excel Cell')
Red = PatternFill(patternType='solid', fgColor='35FC03')
Green = PatternFill(patternType='solid', fgColor='FC2C03')

if returnvalue[0][1] == MurMax:
    ws[row1+str(col)].fill = Green
if returnvalue[1][1] == MurMax:
    ws[row1+str(col+1)].fill = Green
if returnvalue[2][1] == MurMax:
    ws[row1+str(col+2)].fill = Green
if returnvalue[3][1] == MurMax:
    ws[row1+str(col+3)].fill = Green
if returnvalue[4][1] == MurMax:
    ws[row1+str(col+4)].fill = Green
if returnvalue[5][1] == MurMax:
    ws[row1+str(col+5)].fill = Green
if returnvalue[6][1] == MurMax:
    ws[row1+str(col+1)].fill = Green
if returnvalue[7][1] == MurMax:
    ws[row1+str(col+2)].fill = Green
if returnvalue[8][1] == MurMax:
    ws[row1+str(col+3)].fill = Green
if returnvalue[9][1] == MurMax:
    ws[row1+str(col+4)].fill = Green
if returnvalue[10][1] == MurMax:
    ws[row1+str(col+5)].fill = Green

if returnvalue[0][0] == AFMax:
    ws[row2+str(col)].fill = Green
if returnvalue[1][0] == AFMax:
    ws[row2+str(col+1)].fill = Green
if returnvalue[2][0] == AFMax:
    ws[row2+str(col+2)].fill = Green
if returnvalue[3][0] == AFMax:
    ws[row2+str(col+3)].fill = Green
if returnvalue[4][0] == AFMax:
    ws[row2+str(col+4)].fill= Green
if returnvalue[5][0] == AFMax:
    ws[row2+str(col+5)].fill = Green
if returnvalue[6][0] == AFMax:
    ws[row2+str(col+1)].fill = Green
if returnvalue[7][0] == AFMax:
    ws[row2+str(col+2)].fill = Green
if returnvalue[8][0] == AFMax:
    ws[row2+str(col+3)].fill = Green
if returnvalue[9][0] == AFMax:
    ws[row2+str(col+4)].fill= Green
if returnvalue[10][0] == AFMax:
    ws[row2+str(col+5)].fill = Green



if returnvalue[0][1] == MurMin:
    ws[row1+str(col)].fill = Red
if returnvalue[1][1] == MurMin:
    ws[row1+str(col+1)].fill = Red
if returnvalue[2][1] == MurMin:
    ws[row1+str(col+2)].fill = Red
if returnvalue[3][1] == MurMin:
    ws[row1+str(col+3)].fill = Red
if returnvalue[4][1] == MurMin:
    ws[row1+str(col+4)].fill = Red
if returnvalue[5][1] == MurMin:
    ws[row1+str(col+5)].fill = Red
if returnvalue[6][1] == MurMin:
    ws[row1+str(col+1)].fill = Red
if returnvalue[7][1] == MurMin:
    ws[row1+str(col+2)].fill = Red
if returnvalue[8][1] == MurMin:
    ws[row1+str(col+3)].fill = Red
if returnvalue[9][1] == MurMin:
    ws[row1+str(col+4)].fill = Red
if returnvalue[10][1] == MurMin:
    ws[row1+str(col+5)].fill = Red

if returnvalue[0][0] == AFMin:
    ws[row2+str(col)].fill = Red
if returnvalue[1][0] == AFMin:
    ws[row2+str(col+1)].fill = Red
if returnvalue[2][0] == AFMin:
    ws[row2+str(col+2)].fill = Red
if returnvalue[3][0] == AFMin:
    ws[row2+str(col+3)].fill = Red
if returnvalue[4][0] == AFMin:
    ws[row2+str(col+4)].fill= Red
if returnvalue[5][0] == AFMin:
    ws[row2+str(col+5)].fill = Red
if returnvalue[6][0] == AFMin:
    ws[row2+str(col)].fill = Red
if returnvalue[7][0] == AFMin:
    ws[row2+str(col+1)].fill = Red
if returnvalue[8][0] == AFMin:
    ws[row2+str(col+2)].fill = Red
if returnvalue[9][0] == AFMin:
    ws[row2+str(col+3)].fill = Red
if returnvalue[10][0] == AFMin:
    ws[row2+str(col+4)].fill= Red

#Save To excel
wb.save(ExcelFile)
print('Imported to Excel')'''