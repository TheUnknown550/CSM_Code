import numpy as np
import os
import pandas as pd
import librosa
import librosa.display as display
import glob 
import matplotlib.pyplot as plt
import tensorflow
from tensorflow.keras.models import load_model
import openpyxl as px


def extract_data(file_name):
    # function to load files and extract features
    try:
        # here kaiser_fast is a technique used for faster extraction
        X, sample_rate = librosa.load(file_name, res_type='kaiser_fast')
        # we extract mfcc feature from data
        global mfccs
        mfccs = np.mean(librosa.feature.mfcc(
            y=X, sr=sample_rate, n_mfcc=40).T, axis=0)
    except Exception as e:
        #Error Message
        print("Error encountered while parsing file: ")

    #Reshape data to be in the right matrix
    feature = np.array(mfccs).reshape([-1, 1])
    print(mfccs)
    return feature

af = load_model('AI/AF.h5')#AF AI Path
murmur = load_model('AI/murmur.h5') #murmur AI Path

#variables
mfccs = 0.0
data1 = []
data2 = []
data3 = []
data4 = []
data5 = []
data6 = []
ex = '71'
file_name1 = 'C:/Users/mattc/OneDrive/Documents/Project/CSM/Testing/AITest/A22/Test11.wav' #WAV File Path
file_name2 = 'C:/Users/mattc/OneDrive/Documents/Project/CSM/HeartDiseaseSounds/ReducNoise_Testing/TestFile1.wav'
file_name3 = 'C:/Users/mattc/OneDrive/Documents/Project/CSM/HeartDiseaseSounds/ReducNoise_Testing/TestFile2.wav'
file_name4 = 'C:/Users/mattc/OneDrive/Documents/Project/CSM/HeartDiseaseSounds/ReducNoise_Testing/TestFile3.wav'
file_name5 = 'C:/Users/mattc/OneDrive/Documents/Project/CSM/HeartDiseaseSounds/ReducNoise_Testing/TestFile4.wav'
file_name6 = 'C:/Users/mattc/OneDrive/Documents/Project/CSM/HeartDiseaseSounds/ReducNoise_Testing/TestFile5.wav'

#Extract and read the WAV file for the AI
data1.append(extract_data(file_name1))
data2.append(extract_data(file_name2))
data3.append(extract_data(file_name3))
data4.append(extract_data(file_name4))
data5.append(extract_data(file_name5))
data6.append(extract_data(file_name6))

#Analize Predict and Save the data
af_result = af.predict(np.array(data1))
murmur_result = murmur.predict(np.array(data1))
y1 = af_result[0]
b1 = murmur_result[0]

af_result = af.predict(np.array(data2))
murmur_result = murmur.predict(np.array(data2))
y2 = af_result[0]
b2 = murmur_result[0]

af_result = af.predict(np.array(data3))
murmur_result = murmur.predict(np.array(data3))
y3 = af_result[0]
b3 = murmur_result[0]

af_result = af.predict(np.array(data4))
murmur_result = murmur.predict(np.array(data4))
y4 = af_result[0]
b4 = murmur_result[0]

af_result = af.predict(np.array(data5))
murmur_result = murmur.predict(np.array(data5))
y5 = af_result[0]
b5 = murmur_result[0]

af_result = af.predict(np.array(data6))
murmur_result = murmur.predict(np.array(data6))
y6 = af_result[0]
b6 = murmur_result[0]

#Make it in % instead of 0.
af_return = y1[0]*100
murmur_return = b1[0]*100

#Show the ai confidence[AF,murmur]
returnvalue1 = [af_return, murmur_return]
print('Normal: ',returnvalue1)

af_return = y2[0]*100
murmur_return = b2[0]*100
returnvalue2 = [af_return, murmur_return]
print('50hz.: ',returnvalue2)

af_return = y3[0]*100
murmur_return = b3[0]*100
returnvalue3 = [af_return, murmur_return]
print('100hz.: ',returnvalue3)


af_return = y4[0]*100
murmur_return = b4[0]*100
returnvalue4 = [af_return, murmur_return]
print('200hz.: ',returnvalue4)

af_return = y5[0]*100
murmur_return = b5[0]*100
returnvalue5 = [af_return, murmur_return]
print('300hz.: ',returnvalue5)

af_return = y6[0]*100
murmur_return = b6[0]*100
returnvalue6 = [af_return, murmur_return]
print('400hz.: ',returnvalue6)

#Upload to Excel
wb = px.load_workbook('C:/Users/mattc/OneDrive/Documents/Project/CSM/Testing/AITest/ExprimentResult.xlsx')
ws = wb.active

#Input the value to Excel
ws['O'+ex].value = round(returnvalue1[0],3)
ws['P'+ex].value = round(returnvalue1[1],3)
ws['Q'+ex].value = round(returnvalue2[0],3)
ws['R'+ex].value = round(returnvalue2[1],3)
ws['S'+ex].value = round(returnvalue3[0],3)
ws['T'+ex].value = round(returnvalue3[1],3) 
ws['U'+ex].value = round(returnvalue4[0],3)
ws['V'+ex].value = round(returnvalue4[1],3)
ws['W'+ex].value = round(returnvalue5[0],3)
ws['X'+ex].value = round(returnvalue5[1],3)
ws['Y'+ex].value = round(returnvalue6[0],3)
ws['Z'+ex].value = round(returnvalue6[1],3)


#Save To excel
wb.save('C:/Users/mattc/OneDrive/Documents/Project/CSM/Testing/AITest/ExprimentResult.xlsx')
print('Imported to Excel')

