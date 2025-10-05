import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string
from openpyxl.utils.exceptions import CellCoordinatesException

ExcelFile = 'C:/Users/Matt/Documents/Project/CS-M/Experiments/FilterTests/FilterTestResults.xlsx'
# load the Excel workbook and select the worksheet
wb = openpyxl.load_workbook(ExcelFile)
wsbook = wb.active
ws = wb['Total']

# iterate over each column in the worksheet
for col in range(1, ws.max_column+1):
    # get the column letter for the current column
    col_letter = get_column_letter(col)
    
    # get the maximum and minimum value in the current column
    values = []
    for cell in ws[col_letter]:
        if cell.value is not None:
            value = cell.value
            if isinstance(value, str) and value.startswith('='):
                try:
                    # evaluate the formula for the cell and append the result to the values list
                    row, col_idx = coordinate_from_string(value[1:])
                    col_name = get_column_letter(column_index_from_string(col_letter) + col_idx - 1)
                    value = ws[col_name + str(row)].value
                except CellCoordinatesException:
                    # skip the current cell if the formula references a non-existent cell
                    continue
            try:
                values.append(float(value))
            except ValueError:
                # skip non-numeric values
                pass
    
    # check if there are any numeric values in the current column
    if len(values) == 0:
        continue
    
    max_value = max(values)
    min_value = min(values)
    
    # iterate over each cell in the current column
    for cell in ws[col_letter]:
        # highlight the cell green if it contains the maximum value
        if cell.value == max_value:
            cell.fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
        # highlight the cell red if it contains the minimum value
        elif cell.value == min_value:
            cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

# save the workbook
wb.save(ExcelFile)
