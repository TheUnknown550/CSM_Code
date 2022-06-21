from fileinput import filename
import numpy as np
import os
import math
import pandas as pd
import librosa
import librosa.display as display
import glob 
import matplotlib.pyplot as plt
import tensorflow
import contextlib
import openpyxl as px
import wave
from tensorflow.keras.models import load_model
from openpyxl.styles import PatternFill


def extract_data(file_name):
    # function to load files and extract features
    try:
        # here kaiser_fast is a technique used for faster extraction
        X, sample_rate = librosa.load(file_name, res_type='kaiser_fast')
        # we extract mfcc feature from data
        global mfccs
        mfccs = np.mean(librosa.feature.mfcc(
            y=X, sr=sample_rate, n_mfcc=40).T, axis=0)
    except Exception as e:
        #Error Message
        print("Error encountered while parsing file: ")

    #Reshape data to be in the right matrix
    feature = np.array(mfccs).reshape([-1, 1])
    print(mfccs)
    return feature

# from http://stackoverflow.com/questions/13728392/moving-average-or-running-mean
def running_mean(x, windowSize):
  cumsum = np.cumsum(np.insert(x, 0, 0)) 
  return (cumsum[windowSize:] - cumsum[:-windowSize]) / windowSize

# from http://stackoverflow.com/questions/2226853/interpreting-wav-data/2227174#2227174
def interpret_wav(raw_bytes, n_frames, n_channels, sample_width, interleaved = True):

    if sample_width == 1:
        dtype = np.uint8 # unsigned char
    elif sample_width == 2:
        dtype = np.int16 # signed 2-byte short
    else:
        raise ValueError("Only supports 8 and 16 bit audio formats.")

    channels = np.fromstring(raw_bytes, dtype=dtype)

    if interleaved:
        # channels are interleaved, i.e. sample N of channel M follows sample N of channel M-1 in raw data
        channels.shape = (n_frames, n_channels)
        channels = channels.T
    else:
        # channels are not interleaved. All samples from channel M occur before all samples from channel M-1
        channels.shape = (n_channels, n_frames)

    return channels

af = load_model('AI/Models/AF.h5')#AF AI Path
murmur = load_model('AI/Models/murmur.h5') #murmur AI Path

#variables
mfccs = 0.0
data1 = []
data2 = []
data3 = []
data4 = []
data5 = []
data6 = []
y1 = []
b1 = []
y2 = []
b2 = []
y3 = []
b3 = []
y4 = []
b4 = []
y5 = []
b5 = []
y6 = []
b6 = []
AFList = []
MurList = []
returnvalueAF = [[],[],[],[],[],[],[]]
returnvalueMur = [[],[],[],[],[],[],[]]

ex = 25
Sheet = 'LowPassFilterV3'
file = 'C:/Users/mattc/OneDrive/Documents/Project/CSM/Testing/AITest/A22'#WAV File Path

#Make list for all files
file_name = [file+'/Test1.wav', file+'/Test2.wav', file+'/Test3.wav',
file+'/Test4.wav', file+'/Test5.wav', file+'/Test6.wav', file+'/Test7.wav',
file+'/Test8.wav', file+'/Test9.wav', file+'/Test10.wav', file+'/Test11.wav']

ExcelFiles = 'C:/Users/mattc/OneDrive/Documents/Project/CSM/Testing/AITest/ExprimentResult.xlsx'#Excel File
#Extract and read the WAV file for the AI
data1.append(extract_data(file_name[0]))
data1.append(extract_data(file_name[1]))
data1.append(extract_data(file_name[2]))
data1.append(extract_data(file_name[3]))
data1.append(extract_data(file_name[4]))
data1.append(extract_data(file_name[5]))

#lowpass filter cut off frequency
cutOffFrequency1 = 2.0
cutOffFrequency2 = 4.0
cutOffFrequency3 = 6.0
cutOffFrequency4 = 8.0
cutOffFrequency5 = 10.0

#columns values
col1 = 'B'
col2 = 'C'
col3 = 'D'
col4 = 'E'
col5 = 'F'
col6 = 'G'
col7 = 'H'
col8 = 'I'
col9 = 'J'
col10 = 'K'
col11 = 'L'
col12 = 'M'

#output files
outname1 = 'C:/Users/mattc/OneDrive/Documents/Project/CSM/HeartDiseaseSounds/ReducNoise_Testing/TestFile1.wav'
outname2 = 'C:/Users/mattc/OneDrive/Documents/Project/CSM/HeartDiseaseSounds/ReducNoise_Testing/TestFile2.wav'
outname3 = 'C:/Users/mattc/OneDrive/Documents/Project/CSM/HeartDiseaseSounds/ReducNoise_Testing/TestFile3.wav'
outname4 = 'C:/Users/mattc/OneDrive/Documents/Project/CSM/HeartDiseaseSounds/ReducNoise_Testing/TestFile4.wav'
outname5 = 'C:/Users/mattc/OneDrive/Documents/Project/CSM/HeartDiseaseSounds/ReducNoise_Testing/TestFile5.wav'

def filtering(fname):
    with contextlib.closing(wave.open(fname,'rb')) as spf:
        sampleRate = spf.getframerate()
        ampWidth = spf.getsampwidth()
        nChannels = spf.getnchannels()
        nFrames = spf.getnframes()

        # Extract Raw Audio from multi-channel Wav File
        signal = spf.readframes(nFrames*nChannels)
        spf.close()
        channels = interpret_wav(signal, nFrames, nChannels, ampWidth, True)

        # get window size
        # from http://dsp.stackexchange.com/questions/9966/what-is-the-cut-off-frequency-of-a-moving-average-filter
        #50.0 cutoff
        freqRatio1 = (cutOffFrequency1/sampleRate)
        N = int(math.sqrt(0.196196 + freqRatio1**2)/freqRatio1)

        # Use moviung average (only on first channel)
        filtered = running_mean(channels[0], N).astype(channels.dtype)

        wav_file1 = wave.open(outname1, "w")
        wav_file1.setparams((1, ampWidth, sampleRate, nFrames, spf.getcomptype(), spf.getcompname()))
        wav_file1.writeframes(filtered.tobytes('C'))
        wav_file1.close()


        #100.0 cutoff
        freqRatio2 = (cutOffFrequency2/sampleRate)
        N = int(math.sqrt(0.196196 + freqRatio2**2)/freqRatio2)

        # Use moviung average (only on first channel)
        filtered = running_mean(channels[0], N).astype(channels.dtype)

        wav_file2 = wave.open(outname2, "w")
        wav_file2.setparams((1, ampWidth, sampleRate, nFrames, spf.getcomptype(), spf.getcompname()))
        wav_file2.writeframes(filtered.tobytes('C'))
        wav_file2.close()


        #200.0 cutoff
        freqRatio3 = (cutOffFrequency3/sampleRate)
        N = int(math.sqrt(0.196196 + freqRatio3**2)/freqRatio3)

        # Use moviung average (only on first channel)
        filtered = running_mean(channels[0], N).astype(channels.dtype)

        wav_file3 = wave.open(outname3, "w")
        wav_file3.setparams((1, ampWidth, sampleRate, nFrames, spf.getcomptype(), spf.getcompname()))
        wav_file3.writeframes(filtered.tobytes('C'))
        wav_file3.close()

        #cutoff
        freqRatio4 = (cutOffFrequency4/sampleRate)
        N = int(math.sqrt(0.196196 + freqRatio4**2)/freqRatio4)

        # Use moviung average (only on first channel)
        filtered = running_mean(channels[0], N).astype(channels.dtype)

        wav_file4 = wave.open(outname4, "w")
        wav_file4.setparams((1, ampWidth, sampleRate, nFrames, spf.getcomptype(), spf.getcompname()))
        wav_file4.writeframes(filtered.tobytes('C'))
        wav_file4.close()

        #400.o cutoff
        freqRatio5 = (cutOffFrequency5/sampleRate)
        N = int(math.sqrt(0.196196 + freqRatio5**2)/freqRatio5)

        # Use moviung average (only on first channel)
        filtered = running_mean(channels[0], N).astype(channels.dtype)

        wav_file5 = wave.open(outname5, "w")
        wav_file5.setparams((1, ampWidth, sampleRate, nFrames, spf.getcomptype(), spf.getcompname()))
        wav_file5.writeframes(filtered.tobytes('C'))
        wav_file5.close()

for num in range(len(file_name)):
    filtering(file_name[num])
    data1.append(extract_data(file_name[num]))
    data2 = extract_data(outname1)
    data3 = extract_data(outname2)
    data4 = extract_data(outname3)
    data5 = extract_data(outname4)
    data6 = extract_data(outname5)

    #Analize Predict and Save the data
    af_result = af.predict(np.array(data1[num]))
    murmur_result = murmur.predict(np.array(data1[num]))
    y1.append(af_result[0])
    b1.append(murmur_result[0])

    af_result = af.predict(np.array(data2))
    murmur_result = murmur.predict(np.array(data2))
    y2.append(af_result[0])
    b2.append(murmur_result[0])

    af_result = af.predict(np.array(data3))
    murmur_result = murmur.predict(np.array(data3))
    y3.append(af_result[0])
    b3.append(murmur_result[0])

    af_result = af.predict(np.array(data4))
    murmur_result = murmur.predict(np.array(data4))
    y4.append(af_result[0])
    b4.append(murmur_result[0])

    af_result = af.predict(np.array(data5))
    murmur_result = murmur.predict(np.array(data5))
    y5.append(af_result[0])
    b5.append(murmur_result[0])

    af_result = af.predict(np.array(data6))
    murmur_result = murmur.predict(np.array(data6))
    y6.append(af_result[0])
    b6.append(murmur_result[0])

for index in range(len(returnvalueAF)):
        #Make it in % instead of 0.
        af_return = y1[index][0]*100
        murmur_return = b1[index][0]*100
        #Show the ai confidence[AF,murmur]
        returnvalueAF[index].append(af_return)
        returnvalueMur[index].append(murmur_return)
        print('Normal: ',returnvalueAF[index][0],returnvalueMur[index][0])

        af_return = y2[index][0]*100
        murmur_return = b2[index][0]*100
        returnvalueAF[index].append(af_return)
        returnvalueMur[index].append(murmur_return)
        print('Normal: ',returnvalueAF[index][1],returnvalueMur[index][1])

        af_return = y3[index][0]*100
        murmur_return = b3[index][0]*100
        returnvalueAF[index].append(af_return)
        returnvalueMur[index].append(murmur_return)
        print('Normal: ',returnvalueAF[index][2],returnvalueMur[index][2])


        af_return = y4[index][0]*100
        murmur_return = b4[index][0]*100
        returnvalueAF[index].append(af_return)
        returnvalueMur[index].append(murmur_return)
        print('Normal: ',returnvalueAF[index][3],returnvalueMur[index][3])

        af_return = y5[index][0]*100
        murmur_return = b5[index][0]*100
        returnvalueAF[index].append(af_return)
        returnvalueMur[index].append(murmur_return)
        print('Normal: ',returnvalueAF[index][4],returnvalueMur[index][4])

        af_return = y6[index][0]*100
        murmur_return = b6[index][0]*100
        returnvalueAF[index].append(af_return)
        returnvalueMur[index].append(murmur_return)
        print('Normal: ',returnvalueAF[index][5],returnvalueMur[index][5])


#Upload to Excel
wb = px.load_workbook(ExcelFiles)
wsbook = wb.active
ws = wb[Sheet]

#Input the value to Excel
print('Changing Value of Excel Cell')
for n in range(len(returnvalueAF)):
    ws[col1+str(ex+n)].value = round(returnvalueAF[n][0],3)
    ws[col2+str(ex+n)].value = round(returnvalueMur[n][0],3)
    ws[col3+str(ex+n)].value = round(returnvalueAF[n][1],3)
    ws[col4+str(ex+n)].value = round(returnvalueMur[n][1],3)
    ws[col5+str(ex+n)].value = round(returnvalueAF[n][2],3)
    ws[col6+str(ex+n)].value = round(returnvalueMur[n][2],3)
    ws[col7+str(ex+n)].value = round(returnvalueAF[n][3],3)
    ws[col8+str(ex+n)].value = round(returnvalueMur[n][3],3)
    ws[col9+str(ex+n)].value = round(returnvalueAF[n][4],3)
    ws[col10+str(ex+n)].value = round(returnvalueMur[n][4],3)
    ws[col11+str(ex+n)].value = round(returnvalueAF[n][5],3)
    ws[col12+str(ex+n)].value = round(returnvalueMur[n][5],3)
print('Value of Excel Cell changed')

for ran in range(len(file_name)):
    CopiedValueAF = returnvalueAF.copy()
    CopiedValueMur = returnvalueMur.copy()
    #Find min value in AF and Murmur
    CopiedValueAF.sort()
    CopiedValueMur.sort()
    AFMin = CopiedValueAF[0]
    MurMin = CopiedValueMur[0]
    CopiedValueAF.sort(reverse=True)
    CopiedValueMur.sort(reverse=True)
    AFMax = CopiedValueAF[0]
    MurMax = CopiedValueMur[0]
    #Colour the cell in excel
    print('Changing Colour of Excel Cell')
    Red = PatternFill(patternType='solid', fgColor='35FC03')
    Green = PatternFill(patternType='solid', fgColor='FC2C03')
    if returnvalueAF[ran][0] == AFMax:
        ws[col1+ex].fill = Green
    if returnvalueAF[ran][1] == AFMax:
        ws[col3+ex].fill = Green
    if returnvalueAF[ran][2] == AFMax:
        ws[col5+ex].fill = Green
    if returnvalueAF[ran][3] == AFMax:
        ws[col7+ex].fill = Green
    if returnvalueAF[ran][4] == AFMax:
        ws[col9+ex].fill = Green
    if returnvalueAF[ran][5] == AFMax:
        ws[col11+ex].fill = Green

    if returnvalueMur[ran][0] == MurMax:
        ws[col2+ex].fill = Green
    if returnvalueMur[ran][1] == MurMax:
        ws[col4+ex].fill = Green
    if returnvalueMur[ran][2] == MurMax:
        ws[col6+ex].fill = Green
    if returnvalueMur[ran][3] == MurMax:
        ws[col8+ex].fill = Green
    if returnvalueMur[ran][4] == MurMax:
        ws[col10+ex].fill= Green
    if returnvalueMur[ran][5] == MurMax:
        ws[col12+ex].fill = Green

    if returnvalueAF[ran][0] == AFMin:
        ws[col1+ex].fill = Red
    if returnvalueAF[ran][1] == AFMin:
        ws[col3+ex].fill = Red
    if returnvalueAF[ran][2] == AFMin:
        ws[col5+ex].fill = Red
    if returnvalueAF[ran][3] == AFMin:
        ws[col7+ex].fill = Red
    if returnvalueAF[ran][4] == AFMin:
        ws[col9+ex].fill = Red
    if returnvalueAF[ran][5] == AFMin:
        ws[col11+ex].fill = Red

    if returnvalueMur[ran][0] == MurMin:
        ws[col2+ex].fill = Red
    if returnvalueMur[ran][1] == MurMin:
        ws[col4+ex].fill = Red
    if returnvalueMur[ran][2] == MurMin:
        ws[col6+ex].fill = Red
    if returnvalueMur[ran][3] == MurMin:
        ws[col8+ex].fill = Red
    if returnvalueMur[ran][4] == MurMin:
        ws[col10+ex].fill= Red
    if returnvalueMur[ran][5] == MurMin:
        ws[col12+ex].fill = Red

print('Excel colour changed')

#Save To excel
wb.save(ExcelFiles)
print('Imported to Excel')

