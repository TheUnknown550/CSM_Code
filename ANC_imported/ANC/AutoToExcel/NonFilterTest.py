import numpy as np
import matplotlib.pyplot as plt
from scipy.signal import butter, lfilter, freqz
from scipy.io import wavfile
import soundfile as sf
import openpyxl as px
from scipy.signal import firwin, filtfilt
from pydub import AudioSegment
import time
import math


#Varibles
Sheet = 'NoFilter'
dB = [0,0,10,10,20,20,30,30,40,40,50,50]
SNR_col = ['B','E','H','K','N','Q','T','W','Z','AC','AF','AI','AL','AO','AR','AU','AX','BA','BD','BG','BJ','BM','BP','BS','BV','BY','CB','CE','CH','CK','CN','CQ','CT','CW','CZ']
MSE_col = ['C','F','I','L','O','R','U','X','AA','AD','AG','AJ','AM','AP','AS','AV','AY','BB','BE','BH','BK','BN','BQ','BT','BW','BZ','CC','CF','CI','CL','CO','CR','CU','CX','DA']
row = 4
ExcelFile = 'C:/Users/Matt/Documents/Project/CS-M/Experiments/NoiseCancelTests/Results.xlsx'

# Upload to Excel
wb = px.load_workbook(ExcelFile)
wsbook = wb.active
ws = wb[Sheet]

for n in range(14):
    for i in range(40):
        # Trim the Audio
        def trim_wav( originalWavPath, start, end, name ):
            sampleRate, waveData = wavfile.read( originalWavPath )
            startSample = int( start * sampleRate )
            endSample = int( end * sampleRate )
            wavfile.write( name, sampleRate, waveData[startSample:endSample])

        # Trim Audio File to 10 sec
        if n < 7:
            trim_wav("C:/Users/Matt/Documents/Project/CS-M/Experiments/NoiseCancelTests/ControlledVaribles/Heart/normal/("+str(i+1)+").wav", 0,10, 'C:/Users/Matt/Documents/Project/CS-M/Experiments/NoiseCancelTests/TestingAudios/Original.wav')
            trim_wav('C:/Users/Matt/Documents/Project/CS-M/Experiments/NoiseCancelTests/MixedSounds/'+str(n+1)+'00Hz/normal/('+str(i+1)+').wav', 0,10, 'C:/Users/Matt/Documents/Project/CS-M/Experiments/NoiseCancelTests/TestingAudios/MixedFile.wav')
            
        else:
            trim_wav("C:/Users/Matt/Documents/Project/CS-M/Experiments/NoiseCancelTests/ControlledVaribles/Heart/murmur/("+str(i+1)+").wav", 0,10, 'C:/Users/Matt/Documents/Project/CS-M/Experiments/NoiseCancelTests/TestingAudios/Original.wav')
            trim_wav('C:/Users/Matt/Documents/Project/CS-M/Experiments/NoiseCancelTests/MixedSounds/'+str(n-6)+'00Hz/murmur/('+str(i+1)+').wav', 0,10, 'C:/Users/Matt/Documents/Project/CS-M/Experiments/NoiseCancelTests/TestingAudios/MixedFile.wav')
            
        
        # Read in the audio file
        Original, sample_rate = sf.read('C:/Users/Matt/Documents/Project/CS-M/Experiments/NoiseCancelTests/TestingAudios/Original.wav')
        frames, sampleRate = sf.read('C:/Users/Matt/Documents/Project/CS-M/Experiments/NoiseCancelTests/TestingAudios/MixedFile.wav')        


        # Filter Effectiveness Tests
        # Calculate the Noise% (High is good)
        # Trim the ANC array to match the length of the Test array
        frames = frames[:len(Original)]
        # reshape Original to have the same shape as frames
        Original = np.reshape(Original, (Original.shape[0], 1))
        Original = np.tile(Original, (1, frames.shape[1]))
        
        # Calculate SNR
        signal_power = np.mean(np.square(Original))
        noise_power = np.mean(np.square(frames - Original))
        if noise_power == 0:
            SNR = float('inf')  # infinite SNR for zero noise power
        else:
            SNR = 10 * math.log10(signal_power / noise_power)
        print('SNR: ',SNR)

        # Calculate the MSE (low is good)
        RMSE = np.sqrt(np.mean(np.square(Original - frames)))
        print('RMSE: ',RMSE)

        if n >= len(SNR_col):
            print(n)
            print(len(SNR_col))
            print("The value of n has exceeded the number of columns defined in the SNR_col list.")
            break
        
        # Upload to Excel
        print('Low: ',n+1)
        ws[SNR_col[n]+str(row + i)] = SNR
        ws[MSE_col[n]+str(row + i)] = RMSE

        wb.save(ExcelFile)